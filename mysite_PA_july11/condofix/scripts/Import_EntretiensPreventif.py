# -*- coding: utf-8 -*-
"""
CondoFix - Import Entretien Preventif from XLSX into MySQL (preventif).

Features:
- Reads XLSX directly (openpyxl).
- Robust header matching (accent/space normalization).
- Explicit mapping for duplicate month columns (M/J/A repeats).
- Handles "Employé h." vs "Fournisseur h.":
    - If only one is provided -> 1 insert
    - If both provided -> 2 inserts (as per Donald's comment)
- Transactional: commit all or rollback all.
- Dry-run mode.
- DB config via env or CLI.

Example (PowerShell):
  python .\condofix\scripts\Import_EntretiensPreventif.py `
    --client-id 3 `
    --xlsx .\condofix\ImportEntretien2112.xlsx `
    --id-concierge 31 `
    --id-interv-autre 29 `
    --dry-run

Environment variables (recommended):
  CF_DB_HOST, CF_DB_USER, CF_DB_PASSWORD, CF_DB_NAME

SELECT IDIntervenant, IDClient, NomIntervenant
FROM CondoFix$condofix.intervenants
WHERE IDClient = 1
  AND NomIntervenant LIKE '%Concierge%';

  SELECT IDIntervenant, IDClient, NomIntervenant
FROM CondoFix$condofix.intervenants
WHERE IDClient = 1
  AND NomIntervenant LIKE '%Autre%';
"""

from __future__ import annotations

import argparse
import os
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import mysql.connector
from mysql.connector.cursor import MySQLCursor
from openpyxl import load_workbook


# -----------------------------
# DB config
# -----------------------------

@dataclass(frozen=True)
class DbConfig:
    host: str
    user: str
    password: str
    database: str


def _env(name: str, default: Optional[str] = None) -> Optional[str]:
    v = os.getenv(name)
    return v if v not in (None, "") else default


def get_db(cfg: DbConfig):
    return mysql.connector.connect(
        host=cfg.host,
        user=cfg.user,
        password=cfg.password,
        database=cfg.database,
        autocommit=False,
    )


# -----------------------------
# Header normalization / mapping
# -----------------------------

def normalize_header(s: str) -> str:
    s = s.strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))  # remove accents
    s = s.replace("\u00a0", " ")  # non-breaking space
    s = s.replace(".", "")        # remove dots (h. -> h)
    s = " ".join(s.split())       # collapse whitespace
    return s


# CSV/XLSX headers -> internal field names
# NOTE: We map months explicitly including duplicates (M/J/A repeats).
DEFAULT_HEADER_MAP: Dict[str, str] = {
    "element": "ReferenceCarnet",
    "element ": "ReferenceCarnet",
    "élément": "ReferenceCarnet",

    "description": "Description",
    "categorie": "IDCategorie",
    "catégorie": "IDCategorie",

    "frequence ans": "FreqAns",
    "fréquence ans": "FreqAns",

    "date prochain": "DateProchain",

    "employe h": "HresEmploye",
    "employé h": "HresEmploye",
    "fournisseur h": "HresFournisseur",
    "fournisseur h ": "HresFournisseur",

    # Months (preferred disambiguated headers)
    "j": "Janv",
    "f": "Fev",
    "m": "Mars",
    "a": "Avril",
    "m_1": "Mai",
    "j_2": "Juin",
    "j_3": "Juil",
    "a_4": "Aout",
    "s": "Sept",
    "o": "Oct",
    "n": "Nov",
    "d": "Dec",

    # If someone exports with slightly different month labels:
    "janv": "Janv",
    "fev": "Fev",
    "mars": "Mars",
    "avril": "Avril",
    "mai": "Mai",
    "juin": "Juin",
    "juil": "Juil",
    "aout": "Aout",
    "sept": "Sept",
    "oct": "Oct",
    "nov": "Nov",
    "dec": "Dec",
}


# -----------------------------
# Type coercion
# -----------------------------

INT_FIELDS = {"IDCategorie", "FreqAns", "Janv", "Fev", "Mars", "Avril", "Mai", "Juin", "Juil", "Aout", "Sept", "Oct", "Nov", "Dec"}
FLOAT_FIELDS = {"HresEmploye", "HresFournisseur"}
STR_FIELDS = {"ReferenceCarnet", "Description"}
DATE_FIELDS = {"DateProchain"}


def to_int(v: Any) -> int:
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return 0
    try:
        # allow "1", "1.0", etc.
        return int(float(str(v).replace(",", ".").strip()))
    except Exception as e:
        raise ValueError(f"Cannot convert to int: {v!r}") from e


def to_float(v: Any) -> float:
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return 0.0
    try:
        if isinstance(v, str):
            v = v.replace(",", ".").strip()
        return float(v)
    except Exception as e:
        raise ValueError(f"Cannot convert to float: {v!r}") from e


def to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def to_date_str(v: Any) -> Optional[str]:
    """
    Return YYYY-MM-DD or None.
    Accepts:
      - datetime/date objects from openpyxl
      - strings like '2026-01-31', '31/01/2026'
    """
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None

    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()

    s = str(v).strip()
    # try common formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass

    # if Excel gave something numeric or weird -> fail loudly
    raise ValueError(f"Cannot parse DateProchain: {v!r}")


def coerce(field: str, raw: Any) -> Any:
    if field in INT_FIELDS:
        return to_int(raw)
    if field in FLOAT_FIELDS:
        return to_float(raw)
    if field in STR_FIELDS:
        return to_str(raw)
    if field in DATE_FIELDS:
        return to_date_str(raw)
    return raw


# -----------------------------
# XLSX reading
# -----------------------------

def read_xlsx_rows(
    xlsx_path: Path,
    header_map: Dict[str, str],
    sheet_name: Optional[str] = None,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    raw_headers = [cell.value for cell in ws[1]]
    if not raw_headers or all(h is None for h in raw_headers):
        raise ValueError("Header row is empty")

    col_to_field: Dict[int, str] = {}
    warnings: List[str] = []

    for idx, h in enumerate(raw_headers):
        if h is None:
            continue
        nh = normalize_header(str(h))
        field = header_map.get(nh)
        if field is None:
            continue
        col_to_field[idx] = field

    # sanity checks
    required = {"ReferenceCarnet", "Description", "IDCategorie", "FreqAns"}
    missing = required - set(col_to_field.values())
    if missing:
        warnings.append(f"Missing required mapped columns: {sorted(missing)}")

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(v is None or (isinstance(v, str) and v.strip() == "") for v in r):
            continue

        rec: Dict[str, Any] = {}
        for col_idx, field in col_to_field.items():
            raw_val = r[col_idx] if col_idx < len(r) else None
            rec[field] = coerce(field, raw_val)

        # minimal “real row” heuristic
        if not (rec.get("ReferenceCarnet") or rec.get("Description")):
            continue

        rows.append(rec)

    return rows, warnings


# -----------------------------
# DB operations
# -----------------------------

def build_inserts_for_row(
    rec: Dict[str, Any],
    client_id: int,
    id_concierge: int,
    id_interv_autre: int,
    id_type_travail: int,
) -> List[Tuple[Any, ...]]:
    """
    Returns 0..2 insert tuples depending on hours fields.
    """
    # Determine monthly flags (default 0)
    months = {
        "Janv": rec.get("Janv", 0),
        "Fev": rec.get("Fev", 0),
        "Mars": rec.get("Mars", 0),
        "Avril": rec.get("Avril", 0),
        "Mai": rec.get("Mai", 0),
        "Juin": rec.get("Juin", 0),
        "Juil": rec.get("Juil", 0),
        "Aout": rec.get("Aout", 0),
        "Sept": rec.get("Sept", 0),
        "Oct": rec.get("Oct", 0),
        "Nov": rec.get("Nov", 0),
        "Dec": rec.get("Dec", 0),
    }

    # Hours logic: if both provided, insert twice
    h_emp = float(rec.get("HresEmploye", 0.0) or 0.0)
    h_fou = float(rec.get("HresFournisseur", 0.0) or 0.0)

    base = dict(
        IDClient=client_id,
        Description=rec.get("Description", ""),
        IDCategorie=int(rec.get("IDCategorie", 0) or 0),
        ReferenceCarnet=rec.get("ReferenceCarnet", ""),
        DateProchain=rec.get("DateProchain"),  # YYYY-MM-DD or None
        FreqAns=int(rec.get("FreqAns", 0) or 0),
        IDTypeTravail=id_type_travail,
        months=months,
    )

    inserts: List[Tuple[Any, ...]] = []

    def make_tuple(id_intervenant: int, hres: float) -> Tuple[Any, ...]:
        return (
            base["IDClient"],
            base["Description"],
            hres,
            id_intervenant,
            base["IDCategorie"],
            base["ReferenceCarnet"],
            base["DateProchain"],
            base["FreqAns"],
            base["IDTypeTravail"],
            base["months"]["Janv"],
            base["months"]["Fev"],
            base["months"]["Mars"],
            base["months"]["Avril"],
            base["months"]["Mai"],
            base["months"]["Juin"],
            base["months"]["Juil"],
            base["months"]["Aout"],
            base["months"]["Sept"],
            base["months"]["Oct"],
            base["months"]["Nov"],
            base["months"]["Dec"],
        )

    if h_emp > 0 and h_fou > 0:
        inserts.append(make_tuple(id_concierge, h_emp))
        inserts.append(make_tuple(id_interv_autre, h_fou))
    elif h_emp > 0:
        inserts.append(make_tuple(id_concierge, h_emp))
    elif h_fou > 0:
        inserts.append(make_tuple(id_interv_autre, h_fou))
    else:
        # Keep same behavior as Donald's script when no hours: insert with IDIntervenant=0, HresEstimees=0
        inserts.append(make_tuple(0, 0.0))

    return inserts


def insert_preventif_rows(
    cur: MySQLCursor,
    rows: List[Dict[str, Any]],
    client_id: int,
    id_concierge: int,
    id_interv_autre: int,
    id_type_travail: int,
) -> int:
    if not rows:
        return 0

    sql = (
        "INSERT INTO preventif "
        "(IDClient, Description, HresEstimees, IDIntervenant, IDCategorie, ReferenceCarnet, "
        " DateProchain, FreqAns, IDTypeTravail, Janv, Fev, Mars, Avril, Mai, Juin, Juil, Aout, Sept, Oct, Nov, `Dec`) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    )

    values: List[Tuple[Any, ...]] = []
    for rec in rows:
        values.extend(build_inserts_for_row(rec, client_id, id_concierge, id_interv_autre, id_type_travail))

    cur.executemany(sql, values)
    return cur.rowcount


# -----------------------------
# CLI
# -----------------------------

def parse_map_overrides(pairs: List[str]) -> Dict[str, str]:
    overrides: Dict[str, str] = {}
    for p in pairs:
        if "=" not in p:
            raise ValueError(f"Invalid --map value: {p!r} (expected excelHeader=FieldName)")
        left, right = p.split("=", 1)
        overrides[normalize_header(left)] = right.strip()
    return overrides


def main() -> int:
    ap = argparse.ArgumentParser(description="Import Preventif XLSX into preventif table.")
    ap.add_argument("--client-id", type=int, required=True)
    ap.add_argument("--xlsx", type=str, required=True, help="Path to XLSX (relative or absolute)")
    ap.add_argument("--sheet", type=str, default=None, help="Sheet name (default: active sheet)")
    ap.add_argument("--dry-run", action="store_true", help="Rollback at end; prints what would happen")

    # per-client intervenant IDs
    ap.add_argument("--id-concierge", type=int, required=True, help="IDIntervenant for Employé h. (concierge)")
    ap.add_argument("--id-interv-autre", type=int, required=True, help="IDIntervenant for Fournisseur h. (intervenant autre)")
    ap.add_argument("--id-type-travail", type=int, default=3, help="IDTypeTravail to store (default: 3)")

    # header mapping overrides
    ap.add_argument("--map", action="append", default=[], help='Override mapping: "ExcelHeaderText=InternalFieldName" (repeatable)')

    # DB config (env preferred)
    ap.add_argument("--db-host", default=_env("CF_DB_HOST", "localhost"))
    ap.add_argument("--db-user", default=_env("CF_DB_USER"))
    ap.add_argument("--db-password", default=_env("CF_DB_PASSWORD"))
    ap.add_argument("--db-name", default=_env("CF_DB_NAME", "condofix$condofix"))

    args = ap.parse_args()

    if not args.db_user or not args.db_password:
        print("ERROR: Missing DB credentials. Use env CF_DB_USER/CF_DB_PASSWORD or pass --db-user/--db-password.", file=sys.stderr)
        return 2

    xlsx_path = Path(args.xlsx).resolve()

    header_map = dict(DEFAULT_HEADER_MAP)
    header_map.update(parse_map_overrides(args.map))

    rows, warnings = read_xlsx_rows(xlsx_path, header_map, sheet_name=args.sheet)

    print(f"XLSX: {xlsx_path.name}")
    print(f"Detected rows to import: {len(rows)}")
    if warnings:
        print("Warnings:")
        for w in warnings:
            print(f"  - {w}")

    if len(rows) == 0:
        print("Nothing to import. Exiting.")
        return 0

    cfg = DbConfig(host=args.db_host, user=args.db_user, password=args.db_password, database=args.db_name)

    db = get_db(cfg)
    try:
        cur = db.cursor()
        inserted = insert_preventif_rows(
            cur=cur,
            rows=rows,
            client_id=args.client_id,
            id_concierge=args.id_concierge,
            id_interv_autre=args.id_interv_autre,
            id_type_travail=args.id_type_travail,
        )

        if args.dry_run:
            db.rollback()
            print(f"[DRY RUN] Would insert: {inserted} rows (note: may be > source rows if both hours provided).")
        else:
            db.commit()
            print(f"Inserted: {inserted} rows (note: may be > source rows if both hours provided).")

        return 0

    except Exception as e:
        db.rollback()
        print("ERROR: Import failed; transaction rolled back.", file=sys.stderr)
        print(f"Cause: {type(e).__name__}: {e}", file=sys.stderr)
        raise
    finally:
        try:
            db.close()
        except Exception:
            pass


if __name__ == "__main__":
    raise SystemExit(main())
