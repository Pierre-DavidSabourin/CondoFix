# -*- coding: utf-8 -*-
"""
CondoFix - Import Fonds de prévoyance from XLSX into MySQL (fondsprevoyance).

Features:
- Reads XLSX directly (openpyxl).
- Robust header matching (accent/space normalization).
- Optional archiving of existing rows (historique=1) for a client.
- Transactional: commit all or rollback all.
- Dry-run mode.
- Minimal hardcoding: path + client id + DB config via env or CLI.

Example:
  python condofix/scripts/Import_FDP.py --client-id 1 --xlsx condofix/20260103_Urbano_Importation_FDP.xlsx --archive-existing

Environment variables (recommended):
  CF_DB_HOST, CF_DB_USER, CF_DB_PASSWORD, CF_DB_NAME
"""

from __future__ import annotations

import argparse
import os
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import mysql.connector
from mysql.connector.cursor import MySQLCursor
from openpyxl import load_workbook


# -----------------------------
# DB config
# -----------------------------

@dataclass(frozen=True)
class DbConfig:
    host: str
    user: str
    password: str
    database: str


def _env(name: str, default: Optional[str] = None) -> Optional[str]:
    v = os.getenv(name)
    return v if v not in (None, "") else default


def get_db(cfg: DbConfig):
    return mysql.connector.connect(
        host=cfg.host,
        user=cfg.user,
        password=cfg.password,
        database=cfg.database,
        autocommit=False,
    )


# -----------------------------
# Header normalization / mapping
# -----------------------------

def normalize_header(s: str) -> str:
    """
    Normalize headers so that variations in accents/spaces/case don't break mapping.
    """
    s = s.strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))  # remove accents
    s = s.replace("\u00a0", " ")  # non-breaking space
    s = " ".join(s.split())       # collapse whitespace
    return s


# Default mapping: French Excel headers -> DB columns
# You can override via --map "excel_header=db_field"
DEFAULT_HEADER_MAP: Dict[str, str] = {
    "description": "DescriptionDepense",
    "typemtcerempl": "TypeMtceRempl",
    "refgroupe uniformat": "RefGroupeUniformat",
    "refgroupeuniformat": "RefGroupeUniformat",
    "refanalyse": "RefAnalyse",
    "refanalyse ": "RefAnalyse",
    "codeuniformat": "CodeElementUniformat",
    "valeur actuelle de l'intervention": "ValeurActuelleInterv",
    "valeur actuelle de l’intervention": "ValeurActuelleInterv",
    "% assumes par la copropriete": "PartSyndicat",
    "% assumes par la copropriete ": "PartSyndicat",
    "% assumés par la copropriété": "PartSyndicat",
    "annee de la premiere intervention planifiee": "AnProchain",
    "annee de la premiere intervention planifiee ": "AnProchain",
    "année de la première intervention planifiée": "AnProchain",
    "cycle normal d'intervention": "FrequenceAns",
    "inflation 5 ans": "Inflation5ans",
    "inflation 6-15 ans": "Inflation6a15ans",
    "inflation plus 15 ans": "InflationPlus15ans",
    # The column below exists in Excel but NOT in DB. We intentionally ignore it.
    # "valeur actuelle assumee par la copropriete": None,
}


# -----------------------------
# Type coercion
# -----------------------------

INT_FIELDS = {"TypeMtceRempl", "AnProchain", "FrequenceAns"}
FLOAT_FIELDS = {"ValeurActuelleInterv", "PartSyndicat", "Inflation5ans", "Inflation6a15ans", "InflationPlus15ans"}
STR_FIELDS = {"DescriptionDepense", "RefGroupeUniformat", "RefAnalyse", "CodeElementUniformat"}

DB_INSERT_COLUMNS = [
    "IDClient",
    "DescriptionDepense",
    "TypeMtceRempl",
    "RefGroupeUniformat",
    "RefAnalyse",
    "CodeElementUniformat",
    "ValeurActuelleInterv",
    "FrequenceAns",
    "PartSyndicat",
    "AnProchain",
    "Inflation5ans",
    "Inflation6a15ans",
    "InflationPlus15ans",
    "Actif",
    "historique",
]


def to_int(v: Any) -> Optional[int]:
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return int(float(v))
    except Exception as e:
        raise ValueError(f"Cannot convert to int: {v!r}") from e


def to_float(v: Any) -> Optional[float]:
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        if isinstance(v, str):
            v = v.replace(",", ".").strip()
        return float(v)
    except Exception as e:
        raise ValueError(f"Cannot convert to float: {v!r}") from e


def to_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def coerce(db_field: str, raw: Any) -> Any:
    if db_field in INT_FIELDS:
        return to_int(raw)
    if db_field in FLOAT_FIELDS:
        return to_float(raw)
    if db_field in STR_FIELDS:
        return to_str(raw)
    return raw


def normalize_part_syndicat(value: Optional[float]) -> Optional[float]:
    """
    Accept:
    - 1.0 (already ratio)
    - 100 (percent) -> 1.0
    - 50 -> 0.5
    """
    if value is None:
        return None
    if value > 1.0:
        # treat as percent
        return value / 100.0
    return value


# -----------------------------
# XLSX reading
# -----------------------------

def read_xlsx_rows(
    xlsx_path: Path,
    header_map: Dict[str, str],
    sheet_name: Optional[str] = None,
) -> Tuple[List[str], List[Dict[str, Any]], List[str]]:
    """
    Returns:
      - used_db_fields: list of db fields discovered from headers
      - rows: list of dict(db_field -> coerced value)
      - warnings: list of textual warnings for ignored/unmapped headers
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Read header row (row 1)
    raw_headers = [cell.value for cell in ws[1]]
    if not raw_headers or all(h is None for h in raw_headers):
        raise ValueError("Header row is empty")

    # Build column index -> db_field mapping
    col_to_db: Dict[int, str] = {}
    warnings: List[str] = []
    for idx, h in enumerate(raw_headers):
        if h is None:
            continue
        nh = normalize_header(str(h))
        db_field = header_map.get(nh)
        if db_field is None:
            # unmapped or intentionally ignored
            continue
        col_to_db[idx] = db_field

    if "DescriptionDepense" not in col_to_db.values():
        warnings.append("Missing mapped column for Description -> DescriptionDepense")
    if "TypeMtceRempl" not in col_to_db.values():
        warnings.append("Missing mapped column for TypeMtceRempl")
    if "FrequenceAns" not in col_to_db.values():
        warnings.append("Missing mapped column for Cycle normal d'intervention -> FrequenceAns")

    used_db_fields = sorted(set(col_to_db.values()))

    rows: List[Dict[str, Any]] = []
    # Iterate data rows from row 2
    for r in ws.iter_rows(min_row=2, values_only=True):
        # Stop condition: completely empty row
        if r is None or all(v is None or (isinstance(v, str) and v.strip() == "") for v in r):
            continue

        record: Dict[str, Any] = {}
        for col_idx, db_field in col_to_db.items():
            raw_val = r[col_idx] if col_idx < len(r) else None
            record[db_field] = coerce(db_field, raw_val)

        # Minimal “is this a real row?” heuristic:
        # require at least Description OR a code/ref to exist
        if not (record.get("DescriptionDepense") or record.get("CodeElementUniformat") or record.get("RefAnalyse")):
            continue

        # Normalize PartSyndicat if present
        if "PartSyndicat" in record:
            record["PartSyndicat"] = normalize_part_syndicat(record.get("PartSyndicat"))

        rows.append(record)

    return used_db_fields, rows, warnings


# -----------------------------
# DB operations
# -----------------------------

def archive_existing(cur: MySQLCursor, client_id: int, part_syndicat: str = "all") -> int:
    sql = """
        UPDATE fondsprevoyance
        SET historique = 1
        WHERE IDClient = %s
          AND (historique IS NULL OR historique = 0)
    """
    params: List[Any] = [client_id]

    if part_syndicat != "all":
        # PartSyndicat is effectively 0/1 in your data model (owner vs syndicate responsibility)
        sql += " AND PartSyndicat = %s"
        params.append(int(part_syndicat))

    cur.execute(sql, tuple(params))
    return cur.rowcount

def insert_fdp_rows(
    cur: MySQLCursor,
    client_id: int,
    rows: List[Dict[str, Any]],
    historique_new: int,
) -> int:
    if not rows:
        return 0

    # Prepare values in DB_INSERT_COLUMNS order
    values: List[Tuple[Any, ...]] = []
    for rec in rows:
        row_tuple = (
            client_id,
            rec.get("DescriptionDepense"),
            rec.get("TypeMtceRempl"),
            rec.get("RefGroupeUniformat"),
            rec.get("RefAnalyse"),
            rec.get("CodeElementUniformat"),
            rec.get("ValeurActuelleInterv"),
            rec.get("FrequenceAns"),
            rec.get("PartSyndicat"),
            rec.get("AnProchain"),
            rec.get("Inflation5ans"),
            rec.get("Inflation6a15ans"),
            rec.get("InflationPlus15ans"),
            1,  # Actif
            historique_new,
        )
        values.append(row_tuple)

    placeholders = ", ".join(["%s"] * len(DB_INSERT_COLUMNS))
    sql = f"INSERT INTO fondsprevoyance ({', '.join(DB_INSERT_COLUMNS)}) VALUES ({placeholders})"
    cur.executemany(sql, values)
    return cur.rowcount


# -----------------------------
# CLI
# -----------------------------

def parse_map_overrides(pairs: List[str]) -> Dict[str, str]:
    """
    Parse --map entries like: "RéfAnalyse=CodeElementUniformat"
    Left side is Excel header text (we normalize it).
    """
    overrides: Dict[str, str] = {}
    for p in pairs:
        if "=" not in p:
            raise ValueError(f"Invalid --map value: {p!r} (expected excelHeader=dbField)")
        left, right = p.split("=", 1)
        overrides[normalize_header(left)] = right.strip()
    return overrides


def main() -> int:
    ap = argparse.ArgumentParser(description="Import FDP XLSX into fondsprevoyance (optional archive existing).")
    ap.add_argument("--client-id", type=int, required=True)
    ap.add_argument("--xlsx", type=str, required=True, help="Path to XLSX (relative or absolute)")
    ap.add_argument("--sheet", type=str, default=None, help="Sheet name (default: active sheet)")
    ap.add_argument("--archive-existing", action="store_true", help="Set historique=1 for existing rows of this client before insert")
    ap.add_argument("--historique-new", type=int, default=0, choices=[0, 1], help="historique value for newly inserted rows")
    ap.add_argument("--dry-run", action="store_true", help="Rollback at end; prints what would happen")
    ap.add_argument("--max-errors", type=int, default=10, help="Max row errors to show before aborting")
    ap.add_argument("--archive-part-syndicat",type=str,default="all",choices=["all", "0", "1"],help="When --archive-existing is used: archive only rows matching PartSyndicat (0 or 1). Default: all.",
    )

    # Header mapping overrides
    ap.add_argument(
        "--map",
        action="append",
        default=[],
        help='Override mapping: "ExcelHeaderText=DbFieldName" (repeatable)',
    )

    # DB config (env preferred)
    ap.add_argument("--db-host", default=_env("CF_DB_HOST", "localhost"))
    ap.add_argument("--db-user", default=_env("CF_DB_USER"))
    ap.add_argument("--db-password", default=_env("CF_DB_PASSWORD"))
    ap.add_argument("--db-name", default=_env("CF_DB_NAME", "condofix$condofix"))

    args = ap.parse_args()

    if not args.db_user or not args.db_password:
        print("ERROR: Missing DB credentials. Use env CF_DB_USER/CF_DB_PASSWORD or pass --db-user/--db-password.", file=sys.stderr)
        return 2

    xlsx_path = Path(args.xlsx).resolve()

    # Build effective header map
    header_map = dict(DEFAULT_HEADER_MAP)
    header_map.update(parse_map_overrides(args.map))

    # Read XLSX
    used_fields, rows, warnings = read_xlsx_rows(xlsx_path, header_map, sheet_name=args.sheet)

    print(f"XLSX: {xlsx_path.name}")
    print(f"Detected rows to import: {len(rows)}")
    if warnings:
        print("Warnings:")
        for w in warnings:
            print(f"  - {w}")

    if len(rows) == 0:
        print("Nothing to import. Exiting.")
        return 0

    # Basic sanity checks (optional but useful)
    # If PartSyndicat is present, ensure values are in [0, 1.0]
    bad_part = [r.get("PartSyndicat") for r in rows if r.get("PartSyndicat") is not None and not (0.0 <= r["PartSyndicat"] <= 1.0)]
    if bad_part:
        print(f"ERROR: Found PartSyndicat values outside [0,1] after normalization: {bad_part[:5]}", file=sys.stderr)
        return 3

    cfg = DbConfig(host=args.db_host, user=args.db_user, password=args.db_password, database=args.db_name)

    db = get_db(cfg)
    try:
        cur = db.cursor()

        archived_count = 0
        if args.archive_existing:
            archived_count = archive_existing(cur, args.client_id, args.archive_part_syndicat)

        inserted_count = insert_fdp_rows(cur, args.client_id, rows, args.historique_new)

        if args.dry_run:
            db.rollback()
            print(f"[DRY RUN] Would archive: {archived_count} rows; would insert: {inserted_count} rows.")
        else:
            db.commit()
            print(f"Archived: {archived_count} rows; Inserted: {inserted_count} rows.")

        return 0

    except Exception as e:
        db.rollback()
        print("ERROR: Import failed; transaction rolled back.", file=sys.stderr)
        print(f"Cause: {type(e).__name__}: {e}", file=sys.stderr)
        raise
    finally:
        try:
            db.close()
        except Exception:
            pass


if __name__ == "__main__":
    raise SystemExit(main())
